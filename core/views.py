import openpyxl
import pandas as pd
from datetime import date, datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, models
from django.db.models import Sum, Q
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from xhtml2pdf import pisa

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings
from io import BytesIO
import os

from .models import (
    PlanCaptura, ProgramaReal, CursoExcel, Capacitacion, Curso, 
    CargaSTPS, AreaTematica
)
from .forms import (
    CargaSTPSForm, PlanCapturaForm, ProgramaRealForm, 
    CursoExcelForm, CapacitacionForm
)
from docxtpl import DocxTemplate

from .models import (
    PlanCaptura, ProgramaReal, CursoExcel, Capacitacion, Curso, 
    CargaSTPS, AreaTematica, AreaLaboral, SubareaLaboral, CatalogoAgente
)
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum
from django.shortcuts import render, redirect

from .forms import CapacitacionForm
from .models import Capacitacion
# ==========================================
# FUNCIÓN AUXILIAR: ESTILIZAR REPORTES EXCEL
# ==========================================
def generar_excel_estilizado(registros_queryset):
    if registros_queryset.exists():
        df_exportar = pd.DataFrame(list(registros_queryset))
        
        columnas_modelo = [
            'curp', 'nombre', 'primer_apellido', 'segundo_apellido', 'clave_estado', 
            'clave_municipio', 'clave_ocupacion', 'clave_niv_estudio', 'clave_doc_probatorio', 
            'clave_institucion', 'clave_curso', 'nombre_curso', 'clave_area_tematica', 
            'duracion', 'fec_inicio', 'fec_termino', 'clave_tip_agent', 'rfc_agente_stps', 
            'clave_modalidad', 'clave_capacitacion', 'clave_establec'
        ]
        
        df_exportar = df_exportar[[col for col in columnas_modelo if col in df_exportar.columns]]
        
        df_exportar.columns = [
            'CURP', 'NOMBRE', 'PRIMER APELLIDO', 'SEGUNDO APELLIDO', 'CLAVE ESTADO', 
            'CLAVE MUNICIPIO', 'CLAVE OCUPACION', 'CLAVE NIV ESTUDIO', 'CLAVE DOC PROBATORI', 
            'CLAVE INSTITUC', 'CLAVE CURSO', 'NOMBRE CURSO', 'CLAVE AREA TEMATICA', 
            'DURACIO', 'FEC INICIO', 'FEC TERMINO', 'CLAVE TIP AGENT', 'RFC AGENTE STPS', 
            'CLAVE MODALIDAD', 'CLAVE CAPACITACIO', 'CLAVE ESTABLEC'
        ]
    else:
        df_exportar = pd.DataFrame(columns=[
            'CURP', 'NOMBRE', 'PRIMER APELLIDO', 'SEGUNDO APELLIDO', 'CLAVE ESTADO', 
            'CLAVE MUNICIPIO', 'CLAVE OCUPACION', 'CLAVE NIV ESTUDIO', 'CLAVE DOC PROBATORI', 
            'CLAVE INSTITUC', 'CLAVE CURSO', 'NOMBRE CURSO', 'CLAVE AREA TEMATICA', 
            'DURACIO', 'FEC INICIO', 'FEC TERMINO', 'CLAVE TIP AGENT', 'RFC AGENTE STPS', 
            'CLAVE MODALIDAD', 'CLAVE CAPACITACIO', 'CLAVE ESTABLEC'
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Consolidado_STPS.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df_exportar.to_excel(writer, index=False, sheet_name='Base STPS')
        workbook = writer.book
        worksheet = writer.sheets['Base STPS']
        
        fuente_titulo = Font(name='Segoe UI', size=14, bold=True, color='0891B2')
        fuente_cabecera = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        fuente_datos = Font(name='Segoe UI', size=10, color='1E293B')
        
        fill_cabecera = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        fill_cebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        
        alineacion_centro = Alignment(horizontal='center', vertical='center')
        alineacion_izq = Alignment(horizontal='left', vertical='center')
        
        borde_delgado = Border(
            left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
        )
        
        worksheet.insert_rows(1, 2)
        worksheet['A1'] = "REPORTE CONSOLIDADO DE REGISTROS - CATÁLOGO STPS"
        worksheet['A1'].font = fuente_titulo
        worksheet.row_dimensions[1].height = 25
        
        worksheet.row_dimensions[3].height = 28
        for cell in worksheet[3]:
            cell.font = fuente_cabecera
            cell.fill = fill_cabecera
            cell.alignment = alineacion_centro
            cell.border = borde_delgado
            
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=4, max_row=worksheet.max_row), start=4):
            worksheet.row_dimensions[row_idx].height = 20
            for cell in row:
                cell.font = fuente_datos
                cell.border = borde_delgado
                
                if cell.column_letter in ['A', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
                    cell.alignment = alineacion_centro
                else:
                    cell.alignment = alineacion_izq
                
                if row_idx % 2 == 0:
                    cell.fill = fill_cebra
                    
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 14)
                
    return response


# ==========================================
# 1. VISTA DE CARGA MASIVA STPS
# ==========================================
@login_required
@transaction.atomic
def cargar_stps(request):
    descargar_url = None
    columnas = None
    datos_filas = None
    filas_count = 0

    if 'descargar' in request.GET:
        descargar_url = "?descargar=true"

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        accion = request.POST.get('accion') 
        nombre_archivo = excel_file.name.lower()
        
        try:
            if nombre_archivo.endswith('.csv'):
                try:
                    df = pd.read_csv(excel_file, dtype=str, encoding='latin-1', sep=None, engine='python')
                except Exception:
                    excel_file.seek(0)
                    df = pd.read_csv(excel_file, dtype=str, encoding='utf-8', sep=None, engine='python')
            elif nombre_archivo.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(excel_file, dtype=str)
            else:
                messages.error(request, "Formato de archivo no admitido. Sube un archivo .xlsx, .xls o .csv")
                return redirect('cargar_stps')
            
            df.columns = [str(c).strip().upper() for c in df.columns]
            df = df.fillna("")
            
            if accion == 'previsualizar':
                columnas = df.columns.tolist()
                datos_filas = df.head(15).values.tolist()
                filas_count = len(df)
                messages.info(request, "Mostrando una vista previa del formato STPS.")

            elif accion == 'guardar':
                registros_nuevos = 0
                
                if df.empty:
                    messages.error(request, "El archivo está vacío o no se pudo leer ninguna fila.")
                else:
                    for index, row in df.iterrows():
                        def limpiar_dato_stps(nombre_principal, alternativas=[]):
                            val = row.get(nombre_principal, "")
                            if val == "":
                                for alt in alternativas:
                                    val = row.get(alt, "")
                                    if val != "":
                                        break
                            
                            val_str = str(val).strip()
                            if val_str.endswith('.0'):
                                val_str = val_str[:-2]
                            return val_str

                        CargaSTPS.objects.create(
                            curp=limpiar_dato_stps('CURP'),
                            nombre=limpiar_dato_stps('NOMBRE'),
                            primer_apellido=limpiar_dato_stps('PRIMER APELLIDO'),
                            segundo_apellido=limpiar_dato_stps('SEGUNDO APELLIDO'),
                            clave_estado=limpiar_dato_stps('CLAVE ESTADO'),
                            clave_municipio=limpiar_dato_stps('CLAVE MUNICIPIO'),
                            clave_ocupacion=limpiar_dato_stps('CLAVE OCUPACION'),
                            clave_curso=limpiar_dato_stps('CLAVE CURSO'),
                            nombre_curso=limpiar_dato_stps('NOMBRE CURSO'),
                            clave_area_tematica=limpiar_dato_stps('CLAVE AREA TEMATICA'),
                            duracion=limpiar_dato_stps('DURACION', ['DURACIO']), 
                            fec_inicio=limpiar_dato_stps('FEC INICIO', ['FECHA INICIO']),
                            fec_termino=limpiar_dato_stps('FEC TERMINO', ['FECHA TERMINO']),
                            rfc_agente_stps=limpiar_dato_stps('RFC AGENTE STPS'),
                            clave_modalidad=limpiar_dato_stps('CLAVE MODALIDAD'),
                            clave_niv_estudio=limpiar_dato_stps('CLAVE NIV ESTUDIOS', ['CLAVE NIV ESTUDIO']),
                            clave_doc_probatorio=limpiar_dato_stps('CLAVE DOC PROBATORIO', ['CLAVE DOC PROBATORI']),
                            clave_institucion=limpiar_dato_stps('CLAVE INSTITUCION', ['CLAVE INSTITUC']),
                            clave_tip_agent=limpiar_dato_stps('CLAVE TIP AGENTE', ['CLAVE TIP AGENT']),
                            clave_capacitacion=limpiar_dato_stps('CLAVE CAPACITACION', ['CLAVE CAPACITACIO']),
                            clave_establec=limpiar_dato_stps('CLAVE ESTABLECIMIENTO', ['CLAVE ESTABLEC']),
                        )
                        registros_nuevos += 1
                    
                    messages.success(request, f"¡Éxito! Se cargaron {registros_nuevos} registros. Archivo validado correctamente.")
                    return redirect(f"{request.path}?descargar=true")

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo STPS: {e}")

    if 'descargar' in request.GET:
        registros = CargaSTPS.objects.all().values(
            'curp', 'nombre', 'primer_apellido', 'segundo_apellido', 'clave_estado', 
            'clave_municipio', 'clave_ocupacion', 'clave_niv_estudio', 'clave_doc_probatorio', 
            'clave_institucion', 'clave_curso', 'nombre_curso', 'clave_area_tematica', 
            'duracion', 'fec_inicio', 'fec_termino', 'clave_tip_agent', 'rfc_agente_stps', 
            'clave_modalidad', 'clave_capacitacion', 'clave_establec'
        )
        return generar_excel_estilizado(registros)

    return render(request, 'cargar_stps.html', {
        'descargar_url': descargar_url, 
        'columnas': columnas, 
        'datos_filas': datos_filas, 
        'filas_count': filas_count
    })


# ==========================================
# 2. VISTA DE CAPTURA MANUAL STPS
# ==========================================
@login_required
@transaction.atomic
def captura_manual_stps(request):
    if 'descargar' in request.GET:
        registros = CargaSTPS.objects.all().values(
            'curp', 'nombre', 'primer_apellido', 'segundo_apellido', 'clave_estado', 
            'clave_municipio', 'clave_ocupacion', 'clave_niv_estudio', 'clave_doc_probatorio', 
            'clave_institucion', 'clave_curso', 'nombre_curso', 'clave_area_tematica', 
            'duracion', 'fec_inicio', 'fec_termino', 'clave_tip_agent', 'rfc_agente_stps', 
            'clave_modalidad', 'clave_capacitacion', 'clave_establec'
        )
        return generar_excel_estilizado(registros)

    if request.method == 'POST':
        form = CargaSTPSForm(request.POST)
        if form.is_valid():
            nuevo_registro = form.save()
            messages.success(request, f"¡Éxito! Se registró manualmente a {nuevo_registro.nombre} en el catálogo STPS.")
            return redirect('captura_manual_stps')
    else:
        form = CargaSTPSForm()

    return render(request, 'captura_manual_stps.html', {'form': form})


# ==========================================
# 3. VISTAS DE EDICIÓN Y LISTADO STPS
# ==========================================
@login_required
def lista_modificar_stps(request):
    registros = CargaSTPS.objects.all().order_by('-id')
    return render(request, 'lista_modificar_stps.html', {'registros': registros})


@login_required
@transaction.atomic
def editar_registro_stps(request, pk):
    registro = get_object_or_404(CargaSTPS, pk=pk)
    if request.method == 'POST':
        form = CargaSTPSForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            messages.success(request, f"¡Registro de {registro.nombre} actualizado con éxito!")
            return redirect('lista_modificar_stps')
    else:
        form = CargaSTPSForm(instance=registro)
    return render(request, 'editar_registro_stps.html', {'form': form, 'registro': registro})


# ==========================================
# 4. DASHBOARD Y PLAN CAPTURA
# ==========================================
@login_required
def dashboard(request):
    return render(request, 'dashboard.html')


@login_required
def lista_modificar_plan(request):
    registros = PlanCaptura.objects.all()
    return render(request, 'lista_modificar_plan.html', {'registros': registros})


@login_required
@transaction.atomic
def editar_registro_plan(request, pk):
    registro = get_object_or_404(PlanCaptura, pk=pk)
    if request.method == 'POST':
        form = PlanCapturaForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            messages.success(request, "¡Registro actualizado correctamente!")
            return redirect('lista_modificar_plan')
    else:
        form = PlanCapturaForm(instance=registro)
    return render(request, 'editar_registro_plan.html', {'form': form, 'registro': registro})


@login_required
@transaction.atomic
def plan_captura(request):
    form = PlanCapturaForm(request.POST or None)
    registros = PlanCaptura.objects.all()
    if form.is_valid():
        form.save()
        return redirect('plan_captura')
    return render(request, 'plan_captura.html', {'form': form, 'registros': registros})
@login_required
@transaction.atomic
def programa_real(request):
    # ---- Borrar ----
    borrar_id = request.GET.get('borrar_id')
    if borrar_id:
        reg = get_object_or_404(ProgramaReal, id=borrar_id)
        reg.delete()
        messages.success(request, "Registro eliminado correctamente.")
        return redirect('programa_real')

    # ---- Editar / Nuevo ----
    instance_id = request.GET.get('editar_id') or request.POST.get('editar_id')
    instance = None

    if instance_id:
        try:
            instance = ProgramaReal.objects.get(pk=int(instance_id))
        except (ProgramaReal.DoesNotExist, ValueError, TypeError):
            instance = None

    if request.method == 'POST':
        form = ProgramaRealForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                'Registro actualizado correctamente.' if instance else 'Registro guardado correctamente.'
            )
            return redirect('programa_real')
        messages.error(request, 'Revisa los campos del formulario.')
    else:
        form = ProgramaRealForm(instance=instance)

    registros = ProgramaReal.objects.all().order_by('no')
    totales = registros.aggregate(
        total_importe=Sum('importe'),
        total_operativo=Sum('operativo'),
        total_promotores=Sum('promotores'),
        total_administrativo=Sum('administrativo'),
        total_confianza=Sum('confianza'),
        total_constancia=Sum('constancia'),
    )
    total_general_participantes = (
        (totales['total_operativo'] or 0) +
        (totales['total_promotores'] or 0) +
        (totales['total_administrativo'] or 0) +
        (totales['total_confianza'] or 0)
    )

    return render(request, 'programa_real.html', {
        'form': form,
        'registros': registros,
        'totales': totales,
        'total_general_participantes': total_general_participantes,
        'editando': instance is not None,
        'registro_editando': instance,
    })
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum
from django.shortcuts import render, redirect

from .models import ProgramaReal
from .forms import ProgramaRealForm


def limpiar_numero(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return 0
    texto = str(valor).strip().replace(',', '')
    if texto == '' or texto.upper() in ['NAN', 'NONE', 'SIN COSTO', '-']:
        return 0
    try:
        return int(float(texto))
    except (ValueError, TypeError):
        return 0


def limpiar_importe(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return 0.0
    texto = str(valor).strip().upper()
    if texto in ['', 'NAN', 'NONE', 'SIN COSTO', '-']:
        return 0.0
    try:
        return float(str(valor).replace(',', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return 0.0


def limpiar_texto(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ''
    return str(valor).strip()


@login_required
@transaction.atomic
def cargar_programa_real(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        nombre_archivo = excel_file.name.lower()

        if not nombre_archivo.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Sube un archivo Excel (.xlsx o .xls).')
            return redirect('cargar_programa_real')

        try:
            hojas = pd.read_excel(excel_file, sheet_name=None, header=None)

            nombre_hoja = next(
                (h for h in hojas.keys() if 'relacion' in h.lower() or 'relación' in h.lower()),
                list(hojas.keys())[0]
            )
            df_raw = hojas[nombre_hoja]

            # Buscar fila de encabezados
            fila_header = None
            for i, row in df_raw.iterrows():
                valores = [str(c).strip().upper() for c in row.values if c is not None]
                if any('NOMBRE' in v for v in valores) and any(
                    'OPERATIVO' in v or 'CONSTANCIA' in v for v in valores
                ):
                    fila_header = i
                    break

            if fila_header is None:
                messages.error(request, 'No se encontró la fila de encabezados en el Excel.')
                return redirect('cargar_programa_real')

            excel_file.seek(0)
            df = pd.read_excel(excel_file, sheet_name=nombre_hoja, header=fila_header)
            df.columns = [str(c).strip().upper().replace('\n', ' ') for c in df.columns]

            def encontrar_columna(*opciones):
                for opcion in opciones:
                    for col_name in df.columns:
                        if opcion in col_name:
                            return col_name
                return None

            c_no = encontrar_columna('NO.', 'NO ')
            c_nombre = encontrar_columna('NOMBRE')
            c_importe = encontrar_columna('IMPORTE')
            c_tipo = encontrar_columna('TIPO DE ACCIÓN', 'TIPO DE ACCION', 'TIPO')
            c_modalidad = encontrar_columna('MODALIDAD')
            c_fecha = encontrar_columna('FECHA')
            c_instructor = encontrar_columna('INSTRUCTOR')
            c_constancia = encontrar_columna('CONSTANCIA')
            c_operativo = encontrar_columna('OPERATIVO')
            c_promotores = encontrar_columna('PROMOTORES')
            c_admin = encontrar_columna('ADMINISTRATIVO SINDICALIZADO', 'ADMINISTRATIVO')
            c_confianza = encontrar_columna('CONFIANZA')

            # Fallback por posición (Excel oficial)
            # 0 No, 1 Nombre, 2 Importe, 3 Tipo, 4 Modalidad, 5 Fecha,
            # 6 Instructor, 7 Constancia, 8 Operativo, 9 Promotores,
            # 10 Administrativo, 11 Confianza
            cols = list(df.columns)
            if c_admin is None and len(cols) >= 11:
                c_admin = cols[10]
            if c_operativo is None and len(cols) >= 9:
                c_operativo = cols[8]
            if c_promotores is None and len(cols) >= 10:
                c_promotores = cols[9]
            if c_confianza is None and len(cols) >= 12:
                c_confianza = cols[11]
            if c_constancia is None and len(cols) >= 8:
                c_constancia = cols[7]

            if not c_nombre:
                messages.error(request, 'No se encontró la columna NOMBRE.')
                return redirect('cargar_programa_real')

            borrar_antes = request.POST.get('borrar_antes') == 'on'
            if borrar_antes:
                ProgramaReal.objects.all().delete()

            creados = 0
            for _, row in df.iterrows():
                nombre_curso = limpiar_texto(row.get(c_nombre) if c_nombre else '')
                if not nombre_curso:
                    continue
                if 'TOTAL' in nombre_curso.upper():
                    continue

                no_val = limpiar_numero(row.get(c_no)) if c_no else (creados + 1)
                if no_val == 0:
                    no_val = creados + 1

                ProgramaReal.objects.create(
                    no=no_val,
                    nombre=nombre_curso[:255],
                    importe=limpiar_importe(row.get(c_importe)) if c_importe else 0.0,
                    tipo_accion=limpiar_texto(row.get(c_tipo))[:100] if c_tipo else '',
                    modalidad=limpiar_texto(row.get(c_modalidad))[:100] if c_modalidad else '',
                    fecha=limpiar_texto(row.get(c_fecha))[:100] if c_fecha else '',
                    instructor=limpiar_texto(row.get(c_instructor))[:255] if c_instructor else '',
                    constancia=limpiar_numero(row.get(c_constancia)) if c_constancia else 0,
                    operativo=limpiar_numero(row.get(c_operativo)) if c_operativo else 0,
                    promotores=limpiar_numero(row.get(c_promotores)) if c_promotores else 0,
                    administrativo=limpiar_numero(row.get(c_admin)) if c_admin else 0,
                    confianza=limpiar_numero(row.get(c_confianza)) if c_confianza else 0,
                )
                creados += 1

            messages.success(
                request,
                f'¡Éxito! Se cargaron {creados} registros al Programa Real.'
            )
            return redirect('programa_real')

        except Exception as e:
            messages.error(request, f'Error al procesar el Excel: {e}')
            return redirect('cargar_programa_real')

    return render(request, 'cargar_programa_real.html')


from django.utils import timezone
from datetime import datetime
from django.db.models import Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from .models import CursoExcel
from .forms import CursoExcelForm
@login_required
@transaction.atomic
def cursos_nuevos(request):
    anio_actual = datetime.now().year
    anio_seleccionado = int(request.GET.get('anio') or anio_actual)
    mes_seleccionado = request.GET.get('mes') or str(datetime.now().month).zfill(2)

    meses_dict = {
        '01': 'ENERO', '02': 'FEBRERO', '03': 'MARZO', '04': 'ABRIL',
        '05': 'MAYO', '06': 'JUNIO', '07': 'JULIO', '08': 'AGOSTO',
        '09': 'SEPTIEMBRE', '10': 'OCTUBRE', '11': 'NOVIEMBRE', '12': 'DICIEMBRE'
    }
    nombre_mes = meses_dict.get(mes_seleccionado, 'MES NO VÁLIDO')

    instance_id = request.POST.get('editar_id') or request.GET.get('editar_id')
    instance = None
    if instance_id:
        try:
            instance = CursoExcel.objects.get(pk=int(instance_id))
        except (CursoExcel.DoesNotExist, ValueError, TypeError):
            instance = None

    if request.method == 'POST':
        form = CursoExcelForm(request.POST, instance=instance)
        if form.is_valid():
            datos = form.cleaned_data
            nombre_curso = datos['nombre']
            fecha_curso = datos.get('fecha')

            if instance:
                registro = form.save(commit=False)
                registro.fecha_registro = timezone.now().date()
                if fecha_curso:
                    registro.mes = str(fecha_curso.month).zfill(2)
                    registro.anio = fecha_curso.year
                registro.save()
                messages.success(request, "Registro actualizado correctamente.")
            else:
                # Misma fecha + mismo curso → sumar; otra fecha → nueva fila
                registro_existente = None
                if fecha_curso:
                    registro_existente = CursoExcel.objects.filter(
                        nombre=nombre_curso,
                        fecha=fecha_curso
                    ).first()

                if registro_existente:
                    registro_existente.cantidad += datos.get('cantidad') or 0
                    registro_existente.constancia += datos.get('constancia') or 0
                    registro_existente.operativo += datos.get('operativo') or 0
                    registro_existente.promotores += datos.get('promotores') or 0
                    registro_existente.administrativo += datos.get('administrativo') or 0
                    registro_existente.confianza += datos.get('confianza') or 0
                    registro_existente.hombres += datos.get('hombres') or 0
                    registro_existente.mujeres += datos.get('mujeres') or 0
                    registro_existente.fecha_registro = timezone.now().date()
                    registro_existente.save()
                    messages.success(
                        request,
                        f"Se actualizó «{nombre_curso}» del {fecha_curso}: se sumaron participantes."
                    )
                else:
                    nuevo = form.save(commit=False)
                    nuevo.fecha_registro = timezone.now().date()
                    if fecha_curso:
                        nuevo.mes = str(fecha_curso.month).zfill(2)
                        nuevo.anio = fecha_curso.year
                    else:
                        nuevo.mes = mes_seleccionado
                        nuevo.anio = anio_seleccionado

                    ultimo = CursoExcel.objects.filter(
                        mes=nuevo.mes,
                        anio=nuevo.anio
                    ).order_by('-no').first()
                    nuevo.no = (ultimo.no + 1) if ultimo else 1
                    nuevo.save()
                    messages.success(
                        request,
                        f"Curso «{nombre_curso}» ({fecha_curso or 'sin fecha'}) registrado."
                    )

            # Redirigir al mes de la fecha capturada
            if fecha_curso:
                return redirect(
                    f"{request.path}?mes={str(fecha_curso.month).zfill(2)}&anio={fecha_curso.year}"
                )
            return redirect(f"{request.path}?mes={mes_seleccionado}&anio={anio_seleccionado}")
        messages.error(request, "Revisa los campos del formulario.")
    else:
        form = CursoExcelForm(instance=instance)

    registros = CursoExcel.objects.filter(
        mes=mes_seleccionado,
        anio=anio_seleccionado
    ).order_by('fecha', 'no')

    totales = registros.aggregate(
        total_cantidad=Sum('cantidad'),
        total_constancia=Sum('constancia'),
        total_operativo=Sum('operativo'),
        total_promotores=Sum('promotores'),
        total_administrativo=Sum('administrativo'),
        total_confianza=Sum('confianza'),
        total_hombres=Sum('hombres'),
        total_mujeres=Sum('mujeres'),
    )
    total_general = (
        (totales['total_operativo'] or 0) +
        (totales['total_promotores'] or 0) +
        (totales['total_administrativo'] or 0) +
        (totales['total_confianza'] or 0)
    )

    return render(request, 'cursos_nuevos.html', {
        'form': form,
        'registros': registros,
        'totales': totales,
        'total_general': total_general,
        'mes_seleccionado': mes_seleccionado,
        'nombre_mes': nombre_mes,
        'meses_dict': meses_dict,
        'editando': instance is not None,
        'registro_editando': instance,
        'anio_seleccionado': anio_seleccionado,
        'anio_actual': anio_actual,
    })
@login_required
@transaction.atomic
def gestion_capacitaciones(request):
    if request.method == 'POST':
        form = CapacitacionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('gestion_capacitaciones')
    else:
        form = CapacitacionForm()
    registros = Capacitacion.objects.all()
    return render(request, 'capacitaciones.html', {'form': form, 'registros': registros})


@login_required
def buscar_reporte(request):
    query = request.GET.get('q', '').strip()
    registros = PlanCaptura.objects.filter(nomina__icontains=query) if query else PlanCaptura.objects.all()
    return render(request, 'buscar_reporte.html', {'query': query, 'registros': registros})


@login_required
def descargar_pdf(request):
    query = request.GET.get('q', '').strip()
    registros = PlanCaptura.objects.filter(nomina__icontains=query) if query else PlanCaptura.objects.all()
    context = {'registros': registros, 'query': query or 'Todos los registros'}
    template = get_template('pdf_template.html')
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Nomina_{query or "General"}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF', status=500)
    return response


# ==========================================
# CARGA DE EXCEL PLAN CAPTURA
# ==========================================
@login_required
@transaction.atomic
def cargar_excel(request):
    if request.method != "POST" or not request.FILES.get("archivo_excel"):
        return render(request, "core/cargar_excel.html")

    excel_file = request.FILES["archivo_excel"]
    nombre = excel_file.name.lower()
    if not nombre.endswith((".xlsx", ".xls")):
        messages.error(request, "Formato inválido. Sube un archivo Excel (.xlsx o .xls).")
        return redirect("cargar_excel")

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)

        # 1) Elegir la hoja correcta (NO la de nóminas)
        worksheet = None
        for name in wb.sheetnames:
            n = name.upper()
            if "PLAN" in n or "PROGRAMA" in n or "DC-2" in n or "DC2" in n:
                worksheet = wb[name]
                break
        if worksheet is None:
            # fallback: la hoja con más columnas / no NOMINAS
            for name in wb.sheetnames:
                if "NOMINA" not in name.upper():
                    worksheet = wb[name]
                    break
        if worksheet is None:
            worksheet = wb.active

        # 2) Índices por defecto según tu archivo real
        idx_no = 0
        idx_cc = 1
        idx_puesto = 2
        idx_nomina = 3
        idx_nombre = 4
        idx_area = 5
        idx_curp = 6
        idx_induccion = 7
        idx_cap = 10
        idx_modalidad = 13
        idx_curso = 14
        idx_fi = 15
        idx_ft = 16
        idx_duracion = 17
        idx_genero = 18
        idx_costo = 19
        idx_horas = 20

        # 3) Detectar encabezados (filas 1–30)
        header_row = None
        for r in range(1, 31):
            vals = []
            for cell in worksheet[r]:
                vals.append(str(cell.value).strip().upper() if cell.value is not None else "")
            fila_txt = " | ".join(vals)
            if ("NÓMINA" in fila_txt or "NOMINA" in fila_txt) and "CURP" in fila_txt and "PUESTO" in fila_txt:
                header_row = r
                for i, text in enumerate(vals):
                    if text in ("NO", "N°", "Nº") or text.startswith("NO."):
                        idx_no = i
                    elif "CENTRO" in text or text in ("C.C", "CC"):
                        idx_cc = i
                    elif "PUESTO" in text:
                        idx_puesto = i
                    elif "NÓMINA" in text or "NOMINA" in text:
                        idx_nomina = i
                    elif "NOMBRE DEL TRABAJADOR" in text or (text == "NOMBRE") or (
                        "NOMBRE" in text and "CURSO" not in text
                    ):
                        idx_nombre = i
                    elif "ÁREA" in text or "AREA" in text:
                        idx_area = i
                    elif "CURP" in text:
                        idx_curp = i
                    elif "INDUC" in text:
                        idx_induccion = i
                    elif "CAPACIT" in text:
                        idx_cap = i
                    elif "MODALIDAD" in text:
                        idx_modalidad = i
                    elif "NOMBRE DEL CURSO" in text:
                        idx_curso = i
                    elif text == "FI" or "FECHA INICIO" in text or text == "INICIO":
                        idx_fi = i
                    elif text == "FT" or "TÉRMINO" in text or "TERMINO" in text:
                        idx_ft = i
                    elif text == "D" or "DURACION" in text or "DURACIÓN" in text:
                        idx_duracion = i
                    elif "GÉNERO" in text or "GENERO" in text:
                        idx_genero = i
                    elif "COSTO" in text:
                        idx_costo = i
                    elif "TOTAL" in text and "HORA" in text:
                        idx_horas = i
                break

        start_row = (header_row + 2) if header_row else 19  # suele haber subfila FI/FT/D

        def limpio(val):
            if val is None:
                return ""
            s = str(val).strip()
            if s.endswith(".0"):
                try:
                    float(s)
                    s = s[:-2]
                except ValueError:
                    pass
            if s.upper() in ("NONE", "NAN", "#N/A", "#NA", "-"):
                return ""
            return s

        def parsear_fecha(campo):
            if not campo:
                return None
            if isinstance(campo, datetime):
                return campo.date()
            if isinstance(campo, date):
                return campo
            str_f = str(campo).strip()[:10]
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                try:
                    return datetime.strptime(str_f, fmt).date()
                except ValueError:
                    continue
            return None

        def a_entero(val, default=0):
            if val is None or val == "":
                return default
            try:
                return int(float(str(val).replace(",", "").strip()))
            except (ValueError, TypeError):
                return default

        # Valores “arrastrados” (filas hijas de un mismo trabajador)
        ultimo = {
            "no": 1,
            "cc": "",
            "puesto": "",
            "nomina": "",
            "nombre": "",
            "area": "",
            "curp": "",
            "genero": "",
        }

        registros_creados = 0
        registros_omitidos = 0

        for idx_fila, row_tuple in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if idx_fila < start_row:
                continue
            if not row_tuple or all(c is None for c in row_tuple):
                continue

            row = list(row_tuple)
            while len(row) < 25:
                row.append(None)

            val_nombre = limpio(row[idx_nombre] if idx_nombre < len(row) else None)
            val_nomina = limpio(row[idx_nomina] if idx_nomina < len(row) else None)
            val_curso = limpio(row[idx_curso] if idx_curso < len(row) else None)
            val_modalidad = limpio(row[idx_modalidad] if idx_modalidad < len(row) else None)

            # Saltar totales / firmas
            check = f"{val_nombre} {val_nomina} {val_curso}".upper()
            if any(x in check for x in ("TOTAL", "ELABORÓ", "ELABORO", "AUTORIZÓ", "AUTORIZO", "VO.BO")):
                continue

            # Actualizar datos del trabajador si vienen en la fila
            val_no = limpio(row[idx_no] if idx_no < len(row) else None)
            val_cc = limpio(row[idx_cc] if idx_cc < len(row) else None)
            val_puesto = limpio(row[idx_puesto] if idx_puesto < len(row) else None)
            val_area = limpio(row[idx_area] if idx_area < len(row) else None)
            val_curp = limpio(row[idx_curp] if idx_curp < len(row) else None)
            val_genero = limpio(row[idx_genero] if idx_genero < len(row) else None)

            if val_no:
                try:
                    ultimo["no"] = int(float(val_no))
                except ValueError:
                    pass
            if val_cc:
                ultimo["cc"] = val_cc
            if val_puesto:
                ultimo["puesto"] = val_puesto
            if val_nomina:
                ultimo["nomina"] = val_nomina
            if val_nombre:
                ultimo["nombre"] = val_nombre
            if val_area:
                ultimo["area"] = val_area
            if val_curp:
                ultimo["curp"] = val_curp[:18]
            if val_genero:
                g = val_genero.upper()
                if g in ("H", "M", "F", "HOMBRE", "MUJER"):
                    ultimo["genero"] = "H" if g in ("H", "HOMBRE") else "M"

            # Solo guardar si hay CURSO (evita filas vacías de personal sin capacitación)
            if not val_curso:
                registros_omitidos += 1
                continue

            if not ultimo["nombre"] and not ultimo["nomina"]:
                registros_omitidos += 1
                continue

            fi = parsear_fecha(row[idx_fi] if idx_fi < len(row) else None)
            ft = parsear_fecha(row[idx_ft] if idx_ft < len(row) else None)
            if fi and not ft:
                ft = fi
            if not fi:
                fi = date.today()
            if not ft:
                ft = fi

            duracion = a_entero(row[idx_duracion] if idx_duracion < len(row) else None, 0)
            horas = a_entero(row[idx_horas] if idx_horas < len(row) else None, duracion)
            costo = limpio(row[idx_costo] if idx_costo < len(row) else None) or "S/COSTO"

            ind_raw = limpio(row[idx_induccion] if idx_induccion < len(row) else None).upper()
            es_induccion = ind_raw in ("X", "SI", "SÍ", "1", "TRUE")

            cap_raw = limpio(row[idx_cap] if idx_cap < len(row) else None).upper()
            es_capacitacion = cap_raw in ("X", "SI", "SÍ", "1", "TRUE") or bool(val_curso)

            curso_obj, _ = Curso.objects.get_or_create(nombre=val_curso[:255])

            PlanCaptura.objects.create(
                no=ultimo["no"],
                centro_costos=ultimo["cc"][:100] or "S/D",
                puesto=ultimo["puesto"][:100] or "S/D",
                nomina=ultimo["nomina"][:50] or "S/D",
                nombre_trabajador=(ultimo["nombre"] or "")[:255],
                area=ultimo["area"][:100] or "S/D",
                curp=(ultimo["curp"] or "")[:18],
                induccion=es_induccion,
                columna_ind1="",
                columna_ind2="",
                capacitacion=es_capacitacion,
                columna_cap1="",
                columna_cap2="",
                modalidad_curso=(val_modalidad or "")[:100],
                curso=curso_obj,
                fecha_inicio=fi,
                fecha_termino=ft,
                duracion=duracion,
                genero=(ultimo["genero"] or "H")[:10],
                costo=costo[:20],
                total_horas=horas,
            )
            registros_creados += 1

        messages.success(
            request,
            f"Carga terminada. Se insertaron {registros_creados} registros en Plan Captura "
            f"(omitidos sin curso: {registros_omitidos}). Hoja: {worksheet.title}",
        )
        return redirect("plan_captura")

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f"Error al procesar el Excel: {e}")
        return redirect("cargar_excel")

# ==========================================
# VISOR DC-3 (BÚSQUEDA)
# ==========================================
@login_required
def visor_dc3_excel(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('term'):
        q = request.GET.get('term', '').strip()
        q = q.replace('\t', ' ')
        palabras = [p for p in q.split(' ') if p]

        if palabras:
            filtro = Q()
            for palabra in palabras:
                filtro &= (
                    Q(nombre__icontains=palabra) |
                    Q(primer_apellido__icontains=palabra) |
                    Q(segundo_apellido__icontains=palabra)
                )
            registros = CargaSTPS.objects.filter(filtro)[:10]
        else:
            registros = CargaSTPS.objects.none()

        resultados = []
        for reg in registros:
            nombre = f"{reg.nombre or ''} {reg.primer_apellido or ''} {reg.segundo_apellido or ''}"
            nombre = " ".join(nombre.split())

            resultados.append({
                "id": reg.id,
                "label": f"{nombre} — [{reg.nombre_curso or 'SIN CURSO'}]",
                "value": nombre
            })

        return JsonResponse(resultados, safe=False)

    context = {}
    empleado_id = request.GET.get('empleado_id')

    if empleado_id:
        registro = get_object_or_404(CargaSTPS, id=empleado_id)

        try:
            horas = int(registro.duracion)
        except (ValueError, TypeError):
            horas = 0

        nombre = f"{registro.nombre or ''} {registro.primer_apellido or ''} {registro.segundo_apellido or ''}"
        nombre = " ".join(nombre.split())

        areas = AreaLaboral.objects.prefetch_related('subareas').all()
        agentes = CatalogoAgente.objects.all().order_by('nombre')

        context = {
            "registro": registro,
            "nombre_completo": nombre,
            "error_horas": horas < 0,
            "areas": areas,
            "agentes": agentes,
        }

    return render(request, "capacitacion/visor_dc3_excel.html", context)
# ==========================================
# CARGAR ÁREAS TEMÁTICAS
# ==========================================
@login_required
def cargar_areas_tematicas(request):
    if not request.user.is_staff:
        return HttpResponse("Acceso denegado", status=403)

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        
        try:
            dict_excel = pd.read_excel(excel_file, sheet_name=None)
            nombre_pestana = next((sheet for sheet in dict_excel.keys() if 'area' in sheet.lower() or 'tematica' in sheet.lower()), None)
            
            if not nombre_pestana:
                nombre_pestana = list(dict_excel.keys())[0]
            
            df = dict_excel[nombre_pestana]
            df.columns = [str(col).strip().lower() for col in df.columns]
            
            col_clave = next((c for c in df.columns if 'clave' in c or 'id' in c or 'cod' in c), None)
            col_desc = next((c for c in df.columns if 'desc' in c or 'nombre' in c or 'area' in c), None)
            
            if not col_clave or not col_desc:
                messages.error(request, f"El Excel debe tener columnas de Clave y Descripción.")
                return render(request, 'core/cargar_areas.html')
            
            contador_creados = 0
            
            for _, fila in df.iterrows():
                clave_valor = str(fila[col_clave]).strip()
                descripcion_valor = str(fila[col_desc]).strip()
                
                if pd.isna(fila[col_clave]) or clave_valor.lower() == 'nan' or not clave_valor:
                    continue
                
                if clave_valor.endswith('.0'):
                    clave_valor = clave_valor[:-2]
                
                AreaTematica.objects.update_or_create(
                    clave=clave_valor,
                    defaults={'descripcion': descripcion_valor}
                )
                contador_creados += 1
                
            messages.success(request, f"¡Se sincronizaron {contador_creados} áreas temáticas!")
            return redirect('cargar_areas_tematicas')
            
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            
    return render(request, 'core/cargar_areas.html')


# ==========================================
# DESCARGAR DC-3 RELLENO
# ==========================================
@login_required
def descargar_dc3_relleno(request, empleado_id):
    if not request.user.is_staff:
        return HttpResponse("Acceso denegado", status=403)

    registro = get_object_or_404(CargaSTPS, id=empleado_id)

    # =========================
    # 📌 SUBÁREA SELECCIONADA
    # =========================
    texto_area = ""
    texto_subarea = ""
    subarea_id = request.GET.get("subarea_id")

    if subarea_id:
        try:
            subarea = SubareaLaboral.objects.select_related('area').get(id=subarea_id)
            texto_area = f"{subarea.area.codigo} {subarea.area.nombre}".upper()
            texto_subarea = f"{subarea.codigo} {subarea.nombre}".upper()
        except SubareaLaboral.DoesNotExist:
            pass

    # =========================
    # 📌 AGENTE / INSTITUCIÓN SELECCIONADA
    # =========================
    agente_texto = ""
    agente_id = request.GET.get("agente_id")

    if agente_id:
        try:
            agente = CatalogoAgente.objects.get(id=agente_id)
            agente_texto = agente.nombre.upper()
        except CatalogoAgente.DoesNotExist:
            agente_texto = (registro.rfc_agente_stps or "REGISTRO INTERNO LICONSA").upper()
    else:
        agente_texto = (registro.rfc_agente_stps or "REGISTRO INTERNO LICONSA").upper()

    # =========================
    # 🧍 NOMBRE COMPLETO
    # =========================
    nombre_completo = f"{registro.primer_apellido or ''} {registro.segundo_apellido or ''} {registro.nombre or ''}".strip()
    nombre_completo = " ".join(nombre_completo.split()).upper()

    # =========================
    # 📄 PLANTILLA WORD
    # =========================
    ruta_plantilla = os.path.abspath(
        os.path.join(settings.BASE_DIR, "media", "plantillas", "DC-3-Plantilla.docx")
    )

    try:
        doc = DocxTemplate(ruta_plantilla)
    except FileNotFoundError:
        return HttpResponse(f"No se encontró la plantilla en: {ruta_plantilla}", status=404)

    # =========================
    # 🧑‍💼 PUESTO
    # =========================
    puesto_trabajador = ""
    curp_registro = (registro.curp or "").strip()

    if curp_registro:
        plan_empleado = PlanCaptura.objects.filter(curp__iexact=curp_registro).first()
        if plan_empleado:
            puesto_trabajador = getattr(plan_empleado, 'puesto', "")

    if not puesto_trabajador and nombre_completo:
        plan_empleado = PlanCaptura.objects.filter(nombre_trabajador__icontains=nombre_completo).first()
        if plan_empleado:
            puesto_trabajador = getattr(plan_empleado, 'puesto', "")

    puesto_trabajador = (puesto_trabajador or "").upper().strip()

    # =========================
    # 📚 ÁREA TEMÁTICA (del curso)
    # =========================
    area_tematica_curso = ""
    clave_area = str(getattr(registro, 'clave_area_tematica', "") or "").strip()

    if clave_area:
        area_obj = AreaTematica.objects.filter(clave=clave_area).first()
        if area_obj:
            area_tematica_curso = area_obj.descripcion.upper().strip()

    if not area_tematica_curso:
        area_tematica_curso = clave_area.upper()

    # =========================
    # ⏱ HORAS
    # =========================
    valor_duracion = getattr(registro, 'duracion', None) or getattr(registro, 'horas', None) or 0
    try:
        horas_texto = str(int(valor_duracion))
    except (ValueError, TypeError):
        horas_texto = str(valor_duracion).strip()

    # =========================
    # 📅 FECHAS
    # =========================
    def formatear_fecha(valor):
        if not valor:
            return "", "", ""
        if hasattr(valor, "strftime"):
            return valor.strftime("%Y"), valor.strftime("%m"), valor.strftime("%d")
        texto = str(valor).strip().split()[0]
        for formato in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"]:
            try:
                fecha = datetime.strptime(texto, formato)
                return fecha.strftime("%Y"), fecha.strftime("%m"), fecha.strftime("%d")
            except:
                continue
        return "", "", ""

    inicio = getattr(registro, 'fec_inicio', None) or getattr(registro, 'fecha_inicio', None)
    fin = getattr(registro, 'fec_termino', None) or getattr(registro, 'fecha_termino', None)

    ano_ini, mes_ini, dia_ini = formatear_fecha(inicio)
    ano_fin, mes_fin, dia_fin = formatear_fecha(fin)

    # =========================
    # 📦 CONTEXTO WORD
    # =========================
    contexto_word = {
        "nombre": nombre_completo,
        "ocupacion": (registro.clave_ocupacion or "").strip(),
        "empresa": "LECHE PARA EL BIENESTAR LICONSA S.A. DE C.V.",
        "curso": (registro.nombre_curso or "").upper().strip(),
        "agente": agente_texto,
        "puesto": puesto_trabajador,
        "area_tematica": area_tematica_curso,

        # Área y Subárea laboral
        "area_laboral": texto_area,
        "subarea_laboral": texto_subarea,

        "duracion_horas": horas_texto,
        "ano_inicio": ano_ini,
        "mes_inicio": mes_ini,
        "dia_inicio": dia_ini,
        "ano_fin": ano_fin,
        "mes_fin": mes_fin,
        "dia_fin": dia_fin,
    }

    # CURP letra por letra
    curp_texto = (registro.curp or "").upper().strip().ljust(18)
    for i, letra in enumerate(curp_texto):
        contexto_word[f"c{i}"] = letra

    # RFC Empresa
    rfc_texto = "LIC950821M84".upper().strip().ljust(12)
    for i, letra in enumerate(rfc_texto):
        contexto_word[f"r{i}"] = letra

    doc.render(contexto_word)

    output = BytesIO()
    doc.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    filename = f"DC3_{(registro.curp or registro.id).upper().strip()}.docx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum
from django.shortcuts import render, redirect
from .forms import CapacitacionForm
from .models import Capacitacion


def _num_cap(val):
    if val is None:
        return 0
    try:
        if isinstance(val, float) and pd.isna(val):
            return 0
    except Exception:
        pass
    t = str(val).strip().replace(',', '')
    if t == '' or t.upper() in ('NAN', 'NONE', '-'):
        return 0
    try:
        return int(float(t))
    except (ValueError, TypeError):
        return 0


def _txt_cap(val):
    if val is None:
        return ''
    try:
        if isinstance(val, float) and pd.isna(val):
            return ''
    except Exception:
        pass
    return str(val).strip().replace('\n', ' ')


@login_required
@transaction.atomic
def capacitaciones(request):
    instance_id = request.POST.get('editar_id') or request.GET.get('editar_id')
    instance = None
    if instance_id:
        try:
            instance = Capacitacion.objects.get(pk=int(instance_id))
        except (Capacitacion.DoesNotExist, ValueError, TypeError):
            instance = None

    if request.method == 'POST':
        form = CapacitacionForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                'Registro actualizado correctamente.' if instance else 'Registro guardado correctamente.'
            )
            return redirect('capacitaciones')
        messages.error(request, 'Revisa los campos del formulario.')
    else:
        form = CapacitacionForm(instance=instance)

    registros = Capacitacion.objects.all().order_by('consecutivo')
    totales = registros.aggregate(
        total_operativos=Sum('participantes_operativos'),
        total_enlace=Sum('participantes_enlace'),
        total_mandos_medios=Sum('participantes_mandos_medios'),
        total_mandos_superiores=Sum('participantes_mandos_superiores'),
        total_categorias=Sum('participantes_categorias_especiales'),
        total_induccion=Sum('induccion'),
        total_fortalecimiento=Sum('fortalecimiento_desempenio'),
        total_actualizacion=Sum('actualizacion'),
        total_desarrollo=Sum('desarrollo'),
        total_certificacion=Sum('certificacion'),
        total_sensibilizacion=Sum('sensibilizacion'),
        total_hombres=Sum('hombres'),
        total_mujeres=Sum('mujeres'),
    )
    total_participantes = (totales['total_hombres'] or 0) + (totales['total_mujeres'] or 0)

    return render(request, 'capacitaciones.html', {
        'form': form,
        'registros': registros,
        'totales': totales,
        'total_participantes': total_participantes,
        'editando': instance is not None,
        'registro_editando': instance,
    })


@login_required
@transaction.atomic
def cargar_capacitaciones(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        if not archivo.name.lower().endswith(('.xlsx', '.xls', '.xlsm')):
            messages.error(request, 'Sube un archivo Excel (.xlsx / .xls / .xlsm).')
            return redirect('cargar_capacitaciones')

        try:
            # Preferir la hoja del ejemplo 1159
            xl = pd.ExcelFile(archivo)
            hoja = None
            for nombre in xl.sheet_names:
                if 'ejemplo' in nombre.lower() or '1159' in nombre.lower():
                    hoja = nombre
                    break
            if hoja is None:
                hoja = xl.sheet_names[0]

            df = pd.read_excel(archivo, sheet_name=hoja, header=None)

            # Buscar fila con "consecutivo" + "nombre"
            fila_inicio_datos = None
            for i, row in df.iterrows():
                c0 = _txt_cap(row.iloc[0]).lower() if len(row) > 0 else ''
                c1 = _txt_cap(row.iloc[1]).lower() if len(row) > 1 else ''
                if 'consecutivo' in c0 and 'nombre' in c1:
                    # Los datos empiezan 2 filas después (hay subencabezado)
                    fila_inicio_datos = i + 2
                    break

            # Si no encontró encabezado, usar fila 18 (como en tu archivo)
            if fila_inicio_datos is None:
                fila_inicio_datos = 18

            if request.POST.get('borrar_antes') == 'on':
                Capacitacion.objects.all().delete()

            creados = 0
            for i in range(fila_inicio_datos, len(df)):
                row = df.iloc[i]
                nombre = _txt_cap(row.iloc[1]) if len(row) > 1 else ''
                if not nombre:
                    continue
                if nombre.lower() in ('nan', 'total', 'totales'):
                    continue

                Capacitacion.objects.create(
                    consecutivo=_num_cap(row.iloc[0]) if len(row) > 0 else (creados + 1),
                    nombre_tipo_capacitacion=nombre[:255],
                    tipo_capacitacion=_txt_cap(row.iloc[2])[:100] if len(row) > 2 else '',
                    modalidad=_txt_cap(row.iloc[3])[:100] if len(row) > 3 else '',
                    numero_acciones=_num_cap(row.iloc[4]) if len(row) > 4 else 0,
                    participantes_operativos=_num_cap(row.iloc[5]) if len(row) > 5 else 0,
                    participantes_enlace=_num_cap(row.iloc[6]) if len(row) > 6 else 0,
                    participantes_mandos_medios=_num_cap(row.iloc[7]) if len(row) > 7 else 0,
                    participantes_mandos_superiores=_num_cap(row.iloc[8]) if len(row) > 8 else 0,
                    participantes_categorias_especiales=_num_cap(row.iloc[9]) if len(row) > 9 else 0,
                    induccion=_num_cap(row.iloc[11]) if len(row) > 11 else 0,
                    fortalecimiento_desempenio=_num_cap(row.iloc[12]) if len(row) > 12 else 0,
                    actualizacion=_num_cap(row.iloc[13]) if len(row) > 13 else 0,
                    desarrollo=_num_cap(row.iloc[14]) if len(row) > 14 else 0,
                    certificacion=_num_cap(row.iloc[15]) if len(row) > 15 else 0,
                    sensibilizacion=_num_cap(row.iloc[16]) if len(row) > 16 else 0,
                    hombres=_num_cap(row.iloc[23]) if len(row) > 23 else 0,
                    mujeres=_num_cap(row.iloc[24]) if len(row) > 24 else 0,
                )
                creados += 1

            messages.success(request, f'Se cargaron {creados} capacitaciones del Formato 1159.')
            return redirect('capacitaciones')

        except Exception as e:
            messages.error(request, f'Error al procesar el Excel: {e}')
            return redirect('cargar_capacitaciones')

    return render(request, 'cargar_capacitaciones.html')